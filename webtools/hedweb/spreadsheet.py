import os
import io
from flask import current_app
from werkzeug.utils import secure_filename
import openpyxl
from hed import models
from hed import schema as hedschema
from hed.errors.error_reporter import get_printable_issue_string
from hed.errors.exceptions import HedFileError
from hed.validator.event_validator import EventValidator
from hed.util.file_util import get_file_extension

from hedweb.constants import common, file_constants
from hedweb.utils.web_utils import form_has_option, get_hed_schema_from_pull_down
from hedweb.utils.io_utils import generate_filename, get_prefix_dict, file_extension_is_valid, \
    get_text_file_first_row

app_config = current_app.config


def get_columns_info(request):
    columns_file = request.files.get(common.COLUMNS_FILE, '')
    if columns_file:
        filename = columns_file.filename
    else:
        raise HedFileError('MissingSpreadsheetFile', 'An uploadable file was not provided', '')
    worksheet = request.form.get(common.WORKSHEET_SELECTED, None)

    if file_extension_is_valid(filename, file_constants.EXCEL_FILE_EXTENSIONS):
        wb = openpyxl.load_workbook(columns_file, read_only=True)
        worksheet_names = wb.sheetnames
        if not worksheet_names:
            raise HedFileError('BadExcelFile', 'Excel files must worksheets', None)
        elif not worksheet:
            worksheet = worksheet_names[0]
        elif worksheet and worksheet not in worksheet_names:
            raise HedFileError('BadWorksheetName', f'Worksheet {worksheet} not in Excel file', '')
        headers = [c.value for c in next(wb[worksheet].iter_rows(min_row=1, max_row=1))]
        columns_info = {common.COLUMNS_FILE: filename, common.COLUMN_NAMES: headers,
                        common.WORKSHEET_SELECTED: worksheet, common.WORKSHEET_NAMES: worksheet_names}
        wb.close()
    elif file_extension_is_valid(filename, file_constants.TEXT_FILE_EXTENSIONS):
        columns_info = {common.COLUMN_NAMES: get_text_file_first_row(io.StringIO(columns_file.read().decode("utf-8")))}
    else:
        raise HedFileError('BadFileExtension',
                           f'File {filename} extension does not correspond to an Excel or tsv file', '')
    return columns_info


def get_input_from_spreadsheet_form(request):
    """Gets input argument dictionary from the spreadsheet form request.

    Parameters
    ----------
    request: Request object
        A Request object containing user data from the spreadsheet form.

    Returns
    -------
    dictionary
        A dictionary containing input arguments for calling the underlying spreadsheet functions
    """
    arguments = {
        common.SCHEMA: get_hed_schema_from_pull_down(request),
        common.SPREADSHEET: None,
        common.SPREADSHEET_TYPE: file_constants.TSV_EXTENSION,
        common.WORKSHEET_NAME: request.form.get(common.WORKSHEET_SELECTED, None),
        common.COMMAND: request.form.get(common.COMMAND_OPTION, ''),
        common.HAS_COLUMN_NAMES: form_has_option(request, common.HAS_COLUMN_NAMES, 'on'),
        common.CHECK_FOR_WARNINGS: form_has_option(request, common.CHECK_FOR_WARNINGS, 'on'),
    }

    tag_columns, prefix_dict = get_prefix_dict(request.form)
    filename = request.files[common.SPREADSHEET_FILE].filename
    file_ext = get_file_extension(filename)
    if file_ext in file_constants.EXCEL_FILE_EXTENSIONS:
        arguments[common.SPREADSHEET_TYPE] = file_constants.EXCEL_EXTENSION
    spreadsheet = models.HedInput(filename=request.files[common.SPREADSHEET_FILE],
                                  file_type=arguments[common.SPREADSHEET_TYPE],
                                  worksheet_name=arguments.get(common.WORKSHEET_NAME, None),
                                  tag_columns=tag_columns,
                                  has_column_names=arguments.get(common.HAS_COLUMN_NAMES, None),
                                  column_prefix_dictionary=prefix_dict,
                                  display_name=filename)
    arguments[common.SPREADSHEET] = spreadsheet
    return arguments


def spreadsheet_process(arguments):
    """Perform the requested action for the spreadsheet.

    Parameters
    ----------
    arguments: dict
        A dictionary with the input arguments from the spreadsheet form.

    Returns
    -------
      dict
        A dictionary of results from spreadsheet processing in standard form.
    """
    hed_schema = arguments.get('schema', None)
    command = arguments.get(common.COMMAND, None)
    if not hed_schema or not isinstance(hed_schema, hedschema.hed_schema.HedSchema):
        raise HedFileError('BadHedSchema', "Please provide a valid HedSchema", "")
    spreadsheet = arguments.get(common.SPREADSHEET, 'None')
    if not spreadsheet or not isinstance(spreadsheet, models.HedInput):
        raise HedFileError('InvalidSpreadsheet', "An spreadsheet was given but could not be processed", "")

    if command == common.COMMAND_VALIDATE:
        results = spreadsheet_validate(hed_schema, spreadsheet)
    elif command == common.COMMAND_TO_SHORT:
        results = spreadsheet_convert(hed_schema, spreadsheet, command=common.COMMAND_TO_SHORT)
    elif command == common.COMMAND_TO_LONG:
        results = spreadsheet_convert(hed_schema, spreadsheet)
    else:
        raise HedFileError('UnknownSpreadsheetProcessingMethod', f"Command {command} is missing or invalid", "")
    return results


def spreadsheet_convert(hed_schema, spreadsheet, command=common.COMMAND_TO_LONG):
    """Converts a spreadsheet long to short unless unless the command is not COMMAND_TO_LONG then converts to short

    Parameters
    ----------
    hed_schema:HedSchema
        HedSchema object to be used
    spreadsheet: HedInput
        Previously created HedInput object
    command: str
        Name of the command to execute if not COMMAND_TO_LONG

    Returns
    -------
    dict
        A downloadable dictionary file or a file containing warnings
    """

    schema_version = hed_schema.header_attributes.get('version', 'Unknown version')
    results = spreadsheet_validate(hed_schema, spreadsheet)
    if results['data']:
        return results

    display_name = spreadsheet.display_name
    display_ext = os.path.splitext(secure_filename(display_name))[1]

    if command == common.COMMAND_TO_LONG:
        suffix = '_to_long'
        issues = spreadsheet.convert_to_long(hed_schema)
    else:
        suffix = '_to_short'
        issues = spreadsheet.convert_to_short(hed_schema)

    file_name = generate_filename(display_name, suffix=suffix, extension=display_ext)
    # if issues:
    #     issue_str = get_printable_issue_string(issues, f"Spreadsheet {display_name} had conversion errors")
    #     file_name = generate_filename(display_name, suffix='_conversion_errors_' + suffix, extension='.txt')
    #
    #     return {common.COMMAND: command, 'data': issue_str, "output_display_name": file_name,
    #             common.SCHEMA_VERSION: schema_version, "msg_category": "warning",
    #             'msg': f"Spreadsheet {display_name} had conversion errors"}
    # else:
    return {common.COMMAND: command, 'data': '', common.SPREADSHEET: spreadsheet, 'output_display_name': file_name,
            common.SCHEMA_VERSION: schema_version, 'msg_category': 'success',
            'msg': f'Spreadsheet {display_name} converted_successfully'}


def spreadsheet_validate(hed_schema, spreadsheet):
    """ Validates the spreadsheet.

    Parameters
    ----------
    hed_schema: str or HedSchema
        Version number or path or HedSchema object to be used
    spreadsheet: HedFileInput
        Spreadsheet input object to be validated

    Returns
    -------
    dict
         A dictionary containing results of validation in standard format
    """
    schema_version = hed_schema.header_attributes.get('version', 'Unknown version')
    validator = EventValidator(hed_schema=hed_schema)
    issues = validator.validate_input(spreadsheet)
    display_name = spreadsheet.display_name
    if issues:
        issue_str = get_printable_issue_string(issues, f"Spreadsheet {display_name} validation errors")
        file_name = generate_filename(display_name, suffix='_validation_errors', extension='.txt')
        return {common.COMMAND: common.COMMAND_VALIDATE, 'data': issue_str, "output_display_name": file_name,
                common.SCHEMA_VERSION: schema_version, "msg_category": "warning",
                'msg': f"Spreadsheet {display_name} had validation errors"}
    else:
        return {common.COMMAND: common.COMMAND_VALIDATE, 'data': '',
                common.SCHEMA_VERSION: schema_version, 'msg_category': 'success',
                'msg': f'Spreadsheet {display_name} had no validation errors'}
