import json
import os
import tempfile

import openpyxl
from flask import current_app
from hed import models, schema as hedschema
from hed.errors.exceptions import HedFileError
from hed.util.file_util import delete_file_if_it_exists, get_file_extension
from hedweb.constants import file_constants
from werkzeug.utils import secure_filename

import hedweb.constants
from hedweb.constants import common


def get_spreadsheet(arguments):
    if common.SPREADSHEET_STRING in arguments:
        spreadsheet = None
    elif common.SPREADSHEET_PATH in arguments:
        spreadsheet = models.HedInput(arguments.get(common.SPREADSHEET_PATH, None),
                                      worksheet_name=arguments.get(common.WORKSHEET_SELECTED, None),
                                      tag_columns=arguments.get(common.TAG_COLUMNS, None),
                                      has_column_names=arguments.get(common.HAS_COLUMN_NAMES, None),
                                      column_prefix_dictionary=arguments.get(common.COLUMN_PREFIX_DICTIONARY, None))
    else:
        raise HedFileError('NoSpreadsheet', 'No spreadsheet was provided', '')
    return spreadsheet


def delete_file_no_exceptions(file_path):
    try:
        return delete_file_if_it_exists(file_path)
    except:
        return False


def file_extension_is_valid(filename, accepted_file_extensions=None):
    """Checks the other extension against a list of accepted ones.

    Parameters
    ----------
    filename: string
        The name of the other.

    accepted_file_extensions: list
        A list containing all of the accepted other extensions or an empty list of all are accepted

    Returns
    -------
    boolean
        True if the other has a valid other extension.

    """
    if not accepted_file_extensions or os.path.splitext(filename.lower())[1] in accepted_file_extensions:
        return True
    else:
        return False


def file_to_string(file_object):
    """Save a file_object to a string.

    Parameters
    ----------
    file_object: File object
        A other object that points to a other that was first saved in a temporary location.


    Returns
    -------
    string
        The string into which the file was saved.

    """
    filename = ''
    if hasattr(file_object, 'filename'):
        filename = file_object.filename

    file_string = file_object.read()
    return file_string, filename


def generate_filename(base_name, prefix=None, suffix=None, extension=None):
    """Generates a filename for the attachment of the form prefix_basename_suffix + extension.

    Parameters
    ----------
   base_name: str
        The name of the base, usually the name of the file that the issues were generated from
    prefix: str
        The prefix prepended to the front of the base_name
    suffix: str
        The suffix appended to the end of the base_name
    Returns
    -------
    string
        The name of the attachment other containing the issues.
    """

    pieces = []
    if prefix:
        pieces = pieces + [secure_filename(prefix)]
    if base_name:
        pieces.append(os.path.splitext(secure_filename(base_name))[0])
    if suffix:
        pieces = pieces + [secure_filename(suffix)]

    if not pieces:
        return ''
    filename = pieces[0]
    for name in pieces[1:]:
        filename = filename + '_' + name
    if extension:
        filename = filename + '.' + secure_filename(extension)
    return filename


def get_prefix_dict(form_dict):
    """Returns a tag prefix dictionary from a form dictionary.

    Parameters
    ----------
   form_dict: dict
        The dictionary returned from a form that contains a column prefix table
    Returns
    -------
    dict
        A dictionary whose keys are column numbers (starting with 1) and values are tag prefixes to prepend.
    """
    tag_columns = []
    prefix_dict = {}
    keys = form_dict.keys()
    for key in keys:
        if not key.startswith('Column') or key.endswith('check'):
            continue
        pieces = key.split('_')
        check = 'Column_' + pieces[1] + '_check'
        if form_dict.get(check, None) != 'on':
            continue
        if form_dict[key]:
            prefix_dict[int(pieces[1])] = form_dict[key]
        else:
            tag_columns.append(int(pieces[1]))
    return tag_columns, prefix_dict


def get_hed_schema(arguments):
    if common.SCHEMA_STRING in arguments:
        schema_format = arguments.get(common.SCHEMA_FORMAT, ".xml")
        hed_schema = hedschema.from_string(schema_string=arguments[common.SCHEMA_STRING], file_type=schema_format)
    elif common.SCHEMA_PATH in arguments:
        hed_schema = hedschema.load_schema(hed_file_path=arguments[common.SCHEMA_PATH])
    elif common.SCHEMA_URL in arguments:
        # hed_file_path = file_util.url_to_file(arguments[common.SCHEMA_URL])
        hed_schema = hedschema.load_schema(hed_url_path=arguments[common.SCHEMA_URL])
    elif common.SCHEMA_VERSION in arguments:
        hed_file_path = hedschema.get_path_from_hed_version(arguments[common.SCHEMA_VERSION])
        hed_schema = hedschema.load_schema(hed_file_path=hed_file_path)
    else:
        raise HedFileError('NoHEDSchema', 'No valid HED schema was provided', '')
    return hed_schema


def handle_error(ex, hed_info=None, title=None, return_as_str=True):
    """Handles an error by logging and returning a dictionary or simple string

    Parameters
    ----------
    ex: Exception
        The exception raised.
    hed_info: dict
        A dictionary of information.
    title: str
        A title to be included with the message.
    return_as_str: bool
        If true return as string otherwise as dictionary
    Returns
    -------
    str or dict

    """

    if not hed_info:
        hed_info = {}
    if hasattr(ex, 'error_type'):
        error_code = ex.error_type
    else:
        error_code = type(ex).__name__

    if not title:
        title = ''
    if hasattr(ex, 'message'):
        message = ex.message
    else:
        message = str(ex)

    hed_info['message'] = f"{title}[{error_code}: {message}]"
    if return_as_str:
        return json.dumps(hed_info)
    else:
        return hed_info


def save_file_to_upload_folder(file_object, delete_on_close=False):
    """Save a file_object to the upload folder.

    Parameters
    ----------
    file_object: File object
        A other object that points to a other that was first saved in a temporary location.
    delete_on_close: bool
        If true will delete after closing

    Returns
    -------
    string
        The path to the other that was saved to the temporary folder.

    """

    if file_object.filename:
        file_extension = get_file_extension(file_object.filename)
        temporary_upload_file = tempfile.NamedTemporaryFile(suffix=file_extension, delete=delete_on_close,
                                                            dir=current_app.config['UPLOAD_FOLDER'])
        for line in file_object:
            temporary_upload_file.write(line)
        temporary_upload_file.close()
        return temporary_upload_file.name
    else:
        raise("UnableToUploadFile", "File could not uploaded", "UnknownFile")


def save_file_to_upload_folder_no_exception(file_object, delete_on_close=False):
    """Save a other to the upload folder.

    Parameters
    ----------
    file_object: File object
        A other object that points to a other that was first saved in a temporary location.
    delete_on_close: bool
        If true will delete after closing

    Returns
    -------
    string
        The path to the other that was saved to the temporary folder.

    """
    file_path = ''
    try:
        if file_object.filename:
            file_extension = get_file_extension(file_object.filename)
            temporary_upload_file = tempfile.NamedTemporaryFile(suffix=file_extension, delete=delete_on_close,
                                                                dir=current_app.config['UPLOAD_FOLDER'])
            for line in file_object:
                temporary_upload_file.write(line)
            file_path = temporary_upload_file.name
    except:
        file_path = ''
    return file_path


def save_text_to_upload_folder(text, filename):
    """Saves a string in the upload folder as filename.

    Parameters
    ----------
    text: string
        A printable string.
    filename: str
        File name of the issue folder

    Returns
    -------
    string
        The name of the validation output other.

    """

    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    with open(file_path, 'w', encoding='utf-8') as text_file:
        text_file.write(text)
    text_file.close()
    return file_path


def get_text_file_first_row(text_file_path):
    """Gets the contents of the first row of the text file.

    Parameters
    ----------
    text_file_path: string
        The path to a text other.

    Returns
    -------
    list
        list containing first row.

    """
    column_delimiter = get_delimiter_from_file_extension(text_file_path)
    with open(text_file_path, 'r', encoding='utf-8') as opened_text_file:
        first_line = opened_text_file.readline()
        text_file_columns = first_line.split(column_delimiter)
    return text_file_columns


def get_delimiter_from_file_extension(pathname):
    """Gets the file column delimiter based on the extension.

    Parameters
    ----------
    pathname: string
        A path or file name.

    Returns
    -------
    string
        The file column delimiter based on the extension.

    """
    column_delimiter = ''
    file_extension = get_file_extension(pathname).lower()
    if file_extension in hedweb.constants.file_constants.FILE_EXTENSION_TO_DELIMITER_DICTIONARY:
        column_delimiter = hedweb.constants.file_constants.FILE_EXTENSION_TO_DELIMITER_DICTIONARY.get(file_extension)
    return column_delimiter


def get_worksheet_info(file_path, worksheet_name=None):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    worksheet_names = wb.sheetnames
    info = {common.COLUMNS_FILE: file_path, common.WORKSHEET_NAMES: worksheet_names}
    if not worksheet_names:
        raise HedFileError('BadExcelFile', 'Excel files must worksheets', None)
    elif not worksheet_name:
        worksheet_name = worksheet_names[0]
    elif worksheet_name and worksheet_name not in worksheet_names:
        raise HedFileError('BadWorksheetName', f'Worksheet {worksheet_name} not in Excel file', '')
    headers = [c.value for c in next(wb[worksheet_name].iter_rows(min_row=1, max_row=1))]
    info[common.COLUMN_NAMES] = headers
    info[common.WORKSHEET_SELECTED] = worksheet_name
    wb.close()
    return info
