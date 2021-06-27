import openpyxl
from flask import current_app
from hed.errors.exceptions import HedFileError
from hed.util.file_util import get_file_extension
from hedweb.constants import common, file_constants
from hedweb.constants import spreadsheet_constants
from hedweb.utils.io_utils import file_extension_is_valid, save_file_to_upload_folder

app_config = current_app.config


def generate_input_columns_info(request):
    columns_file = request.files.get(common.COLUMNS_FILE, '')
    if columns_file:
        filename = columns_file.filename
    else:
        filename = ''
    columns_path = save_file_to_upload_folder(columns_file)
    input_arguments = {
        common.COLUMNS_PATH: columns_path,
        common.COLUMNS_DISPLAY_NAME: filename,
        common.WORKSHEET_NAME: request.form.get(common.WORKSHEET_SELECTED, None)
    }
    return input_arguments


def get_column_delimiter_based_on_file_extension(file_name_or_path):
    """Gets the spreadsheet column delimiter based on the other extension.

    Parameters
    ----------
    file_name_or_path: string
        A other name or path.

    Returns
    -------
    string
        The spreadsheet column delimiter based on the other extension.

    """
    column_delimiter = ''
    file_extension = get_file_extension(file_name_or_path).lower()
    if file_extension in spreadsheet_constants.SPREADSHEET_FILE_EXTENSION_TO_DELIMITER_DICTIONARY:
        column_delimiter = spreadsheet_constants.SPREADSHEET_FILE_EXTENSION_TO_DELIMITER_DICTIONARY.get(file_extension)
    return column_delimiter


def get_columns_info(input_arguments):
    columns_path = input_arguments.get(common.COLUMNS_PATH)
    worksheet_name = input_arguments.get(common.WORKSHEET_NAME, None)
    columns_display_name = input_arguments.get(common.COLUMNS_DISPLAY_NAME, 'Unknown')
    if file_extension_is_valid(columns_path, file_constants.EXCEL_FILE_EXTENSIONS):
        columns_info = get_worksheet_info(columns_path, worksheet_name)
    elif file_extension_is_valid(columns_path, file_constants.TEXT_FILE_EXTENSIONS):
        columns_info = get_text_file_info(columns_path)
    else:
        raise HedFileError('BadFileExtension',
                           f'File {columns_display_name} extension does not correspond to an Excel or tsv file',
                           columns_display_name)
    return columns_info


def get_other_column_indices(column_names):
    """Gets the other tag column indices in a spreadsheet. The indices found will be one-based.

    Parameters
    ----------
    column_names: list
        A list containing the column names in a spreadsheet.

    Returns
    -------
    list
        A list containing the tag column indices found in a spreadsheet.

    """
    column_indices = []
    for tag_column_name in spreadsheet_constants.OTHER_TAG_COLUMN_NAMES:
        found_indices = find_all_str_indices_in_list(column_names, tag_column_name)
        if found_indices:
            column_indices.extend(found_indices)
    return column_indices


def get_specific_column_indices(column_names):
    """Gets the required tag column indices in a spreadsheet. The indices found will be one-based.

    Parameters
    ----------
    column_names: list
        A list containing the column names in a spreadsheet.

    Returns
    -------
    dictionary
        A dictionary containing the required tag column indices found in a spreadsheet.

    """
    specific_column_indices = {}
    specific_tag_column_names = spreadsheet_constants.SPECIFIC_TAG_COLUMN_NAMES_DICTIONARY.keys()
    for specific_tag_column_name in specific_tag_column_names:
        alternative_specific_tag_column_names = spreadsheet_constants.SPECIFIC_TAG_COLUMN_NAMES_DICTIONARY[
            specific_tag_column_name]
        for alternative_specific_tag_column_name in alternative_specific_tag_column_names:
            indices = find_all_str_indices_in_list(column_names, alternative_specific_tag_column_name)
            if indices:
                specific_column_indices[specific_tag_column_name] = indices[0]
    return specific_column_indices


def get_specific_tag_columns_from_form(request):
    """Gets the specific tag columns from the validation form.

    Parameters
    ----------
    request: Request object
        A Request object containing user data from the validation form.

    Returns
    -------
    dictionary
        A dictionary containing the required tag columns. The keys will be the column numbers and the values will be
        the name of the column.
    """
    column_prefix_dictionary = {}
    for tag_column_name in spreadsheet_constants.SPECIFIC_TAG_COLUMN_NAMES:
        form_tag_column_name = tag_column_name.lower() + common.COLUMN_POSTFIX
        if form_tag_column_name in request.form:
            tag_column_name_index = request.form[form_tag_column_name].strip()
            if tag_column_name_index:
                tag_column_name_index = int(tag_column_name_index)

                # todo: Remove these giant kludges at some point
                if tag_column_name == "Long":
                    tag_column_name = "Long Name"
                tag_column_name = "Event/" + tag_column_name + "/"
                # End giant kludges

                column_prefix_dictionary[tag_column_name_index] = tag_column_name
    return column_prefix_dictionary


def get_text_file_column_names(text_file_path, column_delimiter):
    """Gets the text spreadsheet column names.

    Parameters
    ----------
    text_file_path: string
        The path to a text other.
    column_delimiter: string
        The spreadsheet column delimiter.

    Returns
    -------
    string
        The spreadsheet column delimiter based on the other extension.

    """
    with open(text_file_path, 'r', encoding='utf-8') as opened_text_file:
        first_line = opened_text_file.readline()
        text_file_columns = first_line.split(column_delimiter)
    return text_file_columns


def get_text_file_info(file_path):
    """Populate dictionary with information related to the spreadsheet columns.

    This information contains the names of the spreadsheet columns and column indices that contain HED tags.

    Parameters
    ----------

    file_path: string
        The full path to a spreadsheet other.

    Returns
    -------
    dictionary
        A dictionary populated with information related to the text file columns.

    """

    column_info = {}
    column_delimiter = get_column_delimiter_based_on_file_extension(file_path)
    column_info[common.COLUMN_NAMES] = get_text_file_column_names(file_path, column_delimiter)
    column_info[common.COLUMN_INDICES] = get_other_column_indices(column_info[common.COLUMN_NAMES])
    column_info[common.REQUIRED_COLUMN_INDICES] = get_specific_column_indices(column_info[common.COLUMN_NAMES])
    return column_info


def get_worksheet_info(file_path, worksheet_name=None):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    worksheet_names = wb.sheetnames
    info = {common.COLUMNS_FILE: file_path, common.WORKSHEET_NAMES: worksheet_names}
    if not worksheet_names:
        raise HedFileError('BadExcelFile', 'Excel files must worksheets', None)
    elif not worksheet_name:
        worksheet_name = worksheet_names[0]
    elif worksheet_name and worksheet_name not in worksheet_names:
        raise HedFileError('BadWorksheetName', f'Worksheet {worksheet_name} not in Excel file', '')
    headers = [c.value for c in next(wb[worksheet_name].iter_rows(min_row=1, max_row=1))]
    info[common.COLUMN_NAMES] = headers
    info[common.COLUMN_INDICES] = get_other_column_indices(info[common.COLUMN_NAMES])
    info[common.REQUIRED_COLUMN_INDICES] = get_specific_column_indices(info[common.COLUMN_NAMES])
    info[common.WORKSHEET_SELECTED] = worksheet_name
    wb.close()
    return info


def convert_number_str_to_list(number_str):
    """Converts a string of integers to a list of integers, which is useful for hedweb forms.

    Parameters
    ----------
    number_str: str
        A string containing integers.

    Returns
    -------
    list
        A list containing numbers.
    """
    if number_str:
        return list(map(int, number_str.split(',')))
    return []


def find_all_str_indices_in_list(list_of_str, str_value):
    """Find the indices of a string value in a list.

    Parameters
    ----------
    list_of_str: list
        A list containing strings.
    str_value: string
        A string value.

    Returns
    -------
    list
        A list containing all of the indices where a string value occurs in a string list.

    """
    return [index + 1 for index, value in enumerate(list_of_str) if
            value.lower().replace(' ', '') == str_value.lower().replace(' ', '')]
